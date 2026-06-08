"""生成 正式策划案 GDD Excel — 装备合成系统"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

desktop = os.path.expanduser("~/Desktop")
path = os.path.join(desktop, "装备合成系统_正式策划案_GDD_test.xlsx")

wb = openpyxl.Workbook()

# === Style definitions ===
h1_font = Font(name="Microsoft YaHei", size=16, bold=True, color="2D2418")
h2_font = Font(name="Microsoft YaHei", size=13, bold=True, color="2D2418")
h3_font = Font(name="Microsoft YaHei", size=11, bold=True, color="C4956A")
body_font = Font(name="Microsoft YaHei", size=11, color="2D2418")
small_font = Font(name="Microsoft YaHei", size=10, color="7A6E5E")
header_fill = PatternFill(start_color="FFFBF5", end_color="FFFBF5", fill_type="solid")
accent_fill = PatternFill(start_color="C4956A", end_color="C4956A", fill_type="solid")
p0_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
p1_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
p2_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color="E8DDD0"),
    right=Side(style="thin", color="E8DDD0"),
    top=Side(style="thin", color="E8DDD0"),
    bottom=Side(style="thin", color="E8DDD0"),
)
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="center")

def style_header_row(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="2D2418")
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

def style_data_row(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = body_font
        cell.border = thin_border
        cell.alignment = wrap_align

def auto_width(ws, min_w=12, max_w=50):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = min_w
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_w))
        ws.column_dimensions[letter].width = max_len + 4

# ============================================================
# Sheet 0: §0 文档概述
# ============================================================
ws0 = wb.active
ws0.title = "0-文档概述"

data0 = [
    ["§0 文档概述", "", "", ""],
    ["", "", "", ""],
    ["0.1 游戏名称", "装备合成系统（暂定名：龙鳞工坊）", "", ""],
    ["0.2 一句话介绍", "收集材料→合成装备→强化成长→挑战更强敌人的核心循环，让每一件装备都有\"手造\"的价值感。", "", ""],
    ["", "", "", ""],
    ["0.3 核心循环 (Core Loop)", "战斗/探索 → 收集材料(龙鳞/龙心/矿石...) → 合成工坊合成装备 → 强化提升 → 挑战更强敌人 → 获得稀有材料 → 合成更高级装备", "", ""],
    ["", "", "", ""],
    ["0.4 目标受众与市场定位", "", "", ""],
    ["维度", "描述", "", ""],
    ["目标玩家画像", "18-35岁，喜欢RPG成长感+收集养成，偏好奇幻题材，手机/PC双端", "", ""],
    ["市场定位", "区别于\"一键强化\"的数值堆积，强调材料收集→合成→属性随机的\"手造\"仪式感", "", ""],
    ["竞品分析", "原神圣遗物（随机性太强）/ 怪物猎人装备锻造（确定性太强）→ 本作取中：材料确定产物类型 + 属性范围随机", "", ""],
]

for i, row in enumerate(data0, 1):
    for j, val in enumerate(row, 1):
        ws0.cell(row=i, column=j, value=val)

ws0.merge_cells("A1:D1")
ws0["A1"].font = h1_font
for r in [3,4,6,9]:
    ws0.cell(row=r, column=1).font = h2_font
    ws0.cell(row=r, column=2).font = body_font
for c in range(1,5):
    ws0.cell(row=9, column=c).font = h2_font
for c in range(1,5):
    ws0.cell(row=10, column=c).font = h3_font
    ws0.cell(row=10, column=c).fill = header_fill

auto_width(ws0)

# ============================================================
# Sheet 1: §1 系统设计
# ============================================================
ws1 = wb.create_sheet("1-系统设计")

headers_1 = ["章节", "子系统", "条目", "内容", "关联界面(原型图)"]
for c, h in enumerate(headers_1, 1):
    ws1.cell(row=1, column=c, value=h)
style_header_row(ws1, 1, 5)

sys_data = [
    ["§1.1 核心玩法", "操作机制", "移动/攻击/跳跃/交互", "触屏虚拟摇杆+按钮；手柄左摇杆移动/A攻击/B跳跃/X交互", "01 装备背包"],
    ["§1.1 核心玩法", "战斗系统", "受击判定", "碰撞盒检测，前摇帧不可打断，后摇帧可闪避取消", "战斗HUD"],
    ["§1.1 核心玩法", "战斗系统", "硬直/韧性", "每个敌人有韧性条，累积伤害破韧后进入硬直2s", "战斗HUD"],
    ["§1.1 核心玩法", "战斗系统", "格挡/弹反", "精确时机格挡（0.2s窗口）触发弹反，返还50%伤害", "战斗HUD"],
    ["§1.1 核心玩法", "关卡机制", "地形交互", "可攀爬岩壁（体力消耗）、可破坏木箱（掉落材料）", "-"],
    ["§1.1 核心玩法", "关卡机制", "陷阱/机关", "地刺（周期性弹出）、毒雾区（持续掉血）；压力板开门、火把点燃解锁", "-"],
    ["§1.2 养成局外", "角色成长", "技能树", "角色Lv.10解锁主动技能分支", "角色界面"],
    ["§1.2 养成局外", "角色成长", "装备槽位", "Lv.5/15/30解锁武器/防具/饰品槽", "01 装备背包"],
    ["§1.2 养成局外", "装备系统", "武器", "1个/角色，强化石+金币，每级属性+4%，Lv.80封顶", "06 装备强化"],
    ["§1.2 养成局外", "装备系统", "防具", "1个/角色，强化石+金币，每级防御+3%，Lv.60封顶", "06 装备强化"],
    ["§1.2 养成局外", "装备系统", "饰品", "2个/角色，强化石+金币，每级特殊属性+2%，Lv.50封顶", "06 装备强化"],
    ["§1.2 养成局外", "局内构建", "合成随机性", "配方确定产物类型；属性在范围内随机；稀有材料可锁定1条属性", "04 合成工坊"],
    ["§1.3 辅助系统", "合成图鉴", "记录已解锁配方", "首次合成某配方自动记录", "08 合成图鉴"],
    ["§1.3 辅助系统", "筛选排序", "快速查找装备", "点击筛选按钮弹出弹窗", "09 筛选弹窗"],
    ["§1.3 辅助系统", "容量提示", "背包满时提醒", "容量达到50/50时底栏变红", "01 装备背包"],
]

for i, row in enumerate(sys_data, 2):
    for j, val in enumerate(row, 1):
        ws1.cell(row=i, column=j, value=val)
    style_data_row(ws1, i, 5)

auto_width(ws1, max_w=45)

# ============================================================
# Sheet 2: §2 交互与UI设计
# ============================================================
ws2 = wb.create_sheet("2-交互与UI设计")

# 2.1 界面清单
ws2.cell(row=1, column=1, value="§2.1 界面清单与流程").font = h1_font
ws2.merge_cells("A1:H1")

headers_ui = ["序号", "界面名称", "优先级", "关键元素", "HUD/功能说明", "入口", "状态覆盖", "原型图来源"]
for c, h in enumerate(headers_ui, 1):
    ws2.cell(row=3, column=c, value=h)
style_header_row(ws2, 3, 8)

ui_data = [
    ["01", "装备背包主界面", "P0", "3列网格/卡片/Tab栏/底栏", "浏览所有装备；容量指示；入口枢纽", "主菜单→背包", "默认/空/满/筛选", "Step 3-01"],
    ["02", "装备详情面板", "P0", "3D模型/属性列表/强化+分解按钮", "查看装备完整属性；操作入口", "01 点击装备", "默认/锁定/空槽", "Step 3-02"],
    ["03", "合成入口/配方列表", "P0", "搜索框/分类Tab/配方卡片", "展示所有可合成配方及材料需求", "01 合成工坊按钮", "默认/材料不足/未解锁", "Step 3-03"],
    ["04", "合成工坊", "P0", "预览区/材料槽位(3)/预期属性/开始按钮", "核心合成交互：投放材料→预览→确认", "03 点击配方", "材料未满/已满/合成中", "Step 3-04"],
    ["05", "合成结果弹窗", "P0", "暗色遮罩/武器展示/属性/双按钮", "合成结果反馈：产物属性展示+操作", "04 合成完成", "成功(高品质)/成功(普通)/失败", "Step 3-05"],
    ["06", "装备强化界面", "P1", "属性对比/消耗材料/成功率/强化按钮", "消耗资源提升装备等级", "02 强化按钮", "材料充足/不足/满级", "Step 3-06"],
    ["07", "材料分解界面", "P1", "装备列表+checkbox/预计获得/分解按钮", "分解溢出装备回收材料", "01 分解按钮", "已选/未选/确认弹窗", "Step 3-07"],
    ["08", "合成图鉴", "P2", "进度条/已解锁Tab/未解锁Tab/卡片网格", "收集展示：已解锁配方一览", "主菜单→图鉴", "已解锁/未解锁/空", "Step 3-08"],
    ["09", "筛选/排序弹窗", "P2", "类型checkbox/星级checkbox/排序radio", "快速筛选排序装备列表", "01 筛选按钮", "默认/已选/重置", "Step 3-09"],
]

for i, row in enumerate(ui_data, 4):
    for j, val in enumerate(row, 1):
        ws2.cell(row=i, column=j, value=val)
    style_data_row(ws2, i, 8)
    # Color priority
    prio = row[2]
    fill = p0_fill if prio == "P0" else (p1_fill if prio == "P1" else p2_fill)
    ws2.cell(row=i, column=3).fill = fill

# 2.3 交互反馈
r = 4 + len(ui_data) + 2
ws2.cell(row=r, column=1, value="§2.3 交互反馈设计").font = h1_font
ws2.merge_cells(f"A{r}:H{r}")
r += 2

fb_headers = ["触发事件", "反馈类型", "视觉/听觉/震动", "强度/时长", "关联界面"]
for c, h in enumerate(fb_headers, 1):
    ws2.cell(row=r, column=c, value=h)
style_header_row(ws2, r, 5)
r += 1

fb_data = [
    ["暴击", "视觉+震动", "屏幕震动", "2级 8ms", "战斗HUD"],
    ["合成成功", "视觉+音效+震动", "光柱升起+⭐弹入 / Craft_Success / HD震动", "光柱800ms / stagger 150ms / 震动500ms", "05"],
    ["卡片点击", "视觉", "弹性缩放 0.97→1.0", "100ms spring", "01/03/08"],
    ["按钮hover", "视觉", "上浮2px + 阴影加深", "200ms ease", "全局"],
    ["材料放入", "视觉", "绿色脉冲边框闪烁×2", "400ms", "04"],
    ["UI点击", "听觉", "UI_Click", "按下瞬间", "全局"],
    ["强化成功", "听觉+震动", "Enhance_Success / HD震动", "震动300ms", "06"],
    ["分解确认", "听觉+震动", "Decompose_Confirm / HD震动", "震动150ms", "07"],
]

for row_data in fb_data:
    for c, val in enumerate(row_data, 1):
        ws2.cell(row=r, column=c, value=val)
    style_data_row(ws2, r, 5)
    r += 1

auto_width(ws2, max_w=40)

# ============================================================
# Sheet 3: §3 数值设计
# ============================================================
ws3 = wb.create_sheet("3-数值设计")

ws3.cell(row=1, column=1, value="§3.1 核心战斗公式").font = h1_font
ws3.merge_cells("A1:F1")

ws3.cell(row=3, column=1, value="伤害计算公式").font = h2_font
ws3.cell(row=4, column=1, value="最终伤害 = (攻击力 × 技能倍率 - 防御力 × 0.6) × (1 + 暴击伤害加成) × 元素克制系数").font = Font(name="JetBrains Mono", size=11, bold=True)

dh = ["变量", "说明", "取值范围", "来源"]
for c, h in enumerate(dh, 1):
    ws3.cell(row=6, column=c, value=h)
style_header_row(ws3, 6, 4)

dmg_data = [
    ["攻击力", "角色基础+装备加成", "100~2000", "角色+武器"],
    ["技能倍率", "各技能独立倍率", "0.5~5.0", "技能等级"],
    ["防御力", "敌人属性", "0~800", "敌人等级"],
    ["暴击伤害加成", "基础150%+装备加成", "50%~250%", "武器/饰品"],
    ["元素克制系数", "火>草>水>火 循环", "0.5 / 1.0 / 1.5", "属性相克"],
]
for i, row in enumerate(dmg_data, 7):
    for j, val in enumerate(row, 1):
        ws3.cell(row=i, column=j, value=val)
    style_data_row(ws3, i, 4)

r = 14
ws3.cell(row=r, column=1, value="§3.2 经济系统").font = h1_font
ws3.merge_cells(f"A{r}:F{r}")
r += 2

eh = ["货币", "类型", "获取途径", "主要消耗"]
for c, h in enumerate(eh, 1):
    ws3.cell(row=r, column=c, value=h)
style_header_row(ws3, r, 4)
r += 1

econ_data = [
    ["金币", "软货币", "战斗掉落、分解装备、每日任务", "合成、强化"],
    ["钻石", "硬货币", "充值、成就奖励、活动", "扩充背包、加速合成"],
    ["合成精华", "特殊货币", "分解高品质装备", "锁定属性范围"],
]
for row_data in econ_data:
    for c, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=c, value=val)
    style_data_row(ws3, r, 4)
    r += 1

r += 1
ws3.cell(row=r, column=1, value="§3.3 敌人属性成长曲线").font = h1_font
ws3.merge_cells(f"A{r}:F{r}")
r += 2

gh = ["关卡/波次", "HP 倍率", "攻击力倍率", "移速倍率", "新增行为"]
for c, h in enumerate(gh, 1):
    ws3.cell(row=r, column=c, value=h)
style_header_row(ws3, r, 5)
r += 1

growth_data = [
    ["1-5", "1.0×", "1.0×", "1.0×", "基础攻击"],
    ["6-10", "1.5×", "1.3×", "1.1×", "闪避、格挡"],
    ["11-15", "2.5×", "1.8×", "1.2×", "护盾、召唤小怪"],
    ["16-20", "4.0×", "2.5×", "1.3×", "AOE技能、狂暴(HP<30%)"],
    ["21-25", "6.5×", "3.5×", "1.5×", "双形态、元素免疫循环"],
]
for row_data in growth_data:
    for c, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=c, value=val)
    style_data_row(ws3, r, 5)
    r += 1

auto_width(ws3, max_w=45)

# ============================================================
# Sheet 4: §4 美术需求
# ============================================================
ws4 = wb.create_sheet("4-美术需求")

ws4.cell(row=1, column=1, value="§4.1 视觉风格定调").font = h1_font
ws4.merge_cells("A1:E1")

style_headers = ["维度", "选择", "依据"]
for c, h in enumerate(style_headers, 1):
    ws4.cell(row=3, column=c, value=h)
style_header_row(ws4, 3, 3)

style_data = [
    ["整体风格", "奇幻 RPG 风格化渲染", "原型图 Step 0 结论"],
    ["UI 风格", "高奢柔软 Premium Soft", "装备展示需要质感层次+贵重感"],
    ["配色", "暖奶油底 #FFFBF5 / 深咖啡 #2D2418 / 铜金 #C4956A", "原型图 Step 0"],
    ["字体", "标题:无衬线粗体(700) / 正文:清晰无衬线(400)", "可读性优先"],
]
for i, row in enumerate(style_data, 4):
    for j, val in enumerate(row, 1):
        ws4.cell(row=i, column=j, value=val)
    style_data_row(ws4, i, 3)

r = 9
ws4.cell(row=r, column=1, value="§4.2 角色动画列表（主角）").font = h1_font
ws4.merge_cells(f"A{r}:E{r}")
r += 2

ah = ["动画", "帧数/时长", "是否可打断", "备注"]
for c, h in enumerate(ah, 1):
    ws4.cell(row=r, column=c, value=h)
style_header_row(ws4, r, 4)
r += 1

anim_data = [
    ["待机 Idle", "循环 / 60帧", "-", "含呼吸微动+装备微光"],
    ["跑动 Run", "循环 / 24帧", "是", "起步(4帧)/循环(16帧)/停止(4帧)"],
    ["跳跃 Jump", "30帧 / 1s", "否", "起跳8帧/滞空14帧/落地8帧"],
    ["轻攻击1/2/3", "各12帧 / 0.4s", "后摇可打断", "三连招递增范围"],
    ["重攻击", "30帧 / 1s", "否", "蓄力18帧+释放12帧"],
    ["受击 Hit", "10帧 / 0.33s", "否", "身体后仰+装备闪烁"],
    ["死亡 Death", "40帧 / 1.33s", "否", "倒地+装备散落粒子"],
]
for row_data in anim_data:
    for c, val in enumerate(row_data, 1):
        ws4.cell(row=r, column=c, value=val)
    style_data_row(ws4, r, 4)
    r += 1

r += 1
ws4.cell(row=r, column=1, value="§4.4 特效需求清单").font = h1_font
ws4.merge_cells(f"A{r}:E{r}")
r += 2

vh = ["特效", "触发事件", "类型", "优先级", "关联原型图"]
for c, h in enumerate(vh, 1):
    ws4.cell(row=r, column=c, value=h)
style_header_row(ws4, r, 5)
r += 1

vfx_data = [
    ["合成光柱 VFX", "合成完成", "粒子：金色光柱从底部升起", "P0", "05"],
    ["星级逐个点亮", "合成结果展示", "序列帧：⭐从灰变金 stagger 150ms", "P0", "05"],
    ["材料放入 VFX", "材料拖入槽位", "粒子：绿色脉冲扩散", "P0", "04"],
    ["强化成功 VFX", "强化完成", "粒子：金色上扬+属性绿色飞字", "P1", "06"],
    ["分解 VFX", "装备分解", "粒子：装备碎成金粉+材料图标飞出", "P1", "07"],
    ["按钮 hover VFX", "按钮hover", "缓动：上浮2px+阴影扩散", "P2", "全局"],
    ["God Rays", "合成工坊环境", "粒子：顶部斜射光柱", "P2", "04"],
]
for row_data in vfx_data:
    for c, val in enumerate(row_data, 1):
        ws4.cell(row=r, column=c, value=val)
    style_data_row(ws4, r, 5)
    prio = row_data[3]
    fill = p0_fill if prio == "P0" else (p1_fill if prio == "P1" else p2_fill)
    ws4.cell(row=r, column=4).fill = fill
    r += 1

auto_width(ws4, max_w=45)

# ============================================================
# Save
# ============================================================
wb.save(path)
print(f"✅ GDD Excel 已保存: {path}")
